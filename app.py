from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
from io import BytesIO
import json
from fpdf import FPDF
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = Flask(__name__)

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# Database configuration - supports both SQLite (local) and MySQL (production)
if os.getenv('DATABASE_URL'):
    # Production: Use MySQL from environment variable
    app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
else:
    # Local development: Use SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///mobile_inventory.db'

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-change-in-production')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Simplified Models for new structure
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Brand(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    models = db.relationship('Model', backref='brand', lazy=True)
    incentives = db.relationship('Incentive', backref='brand', lazy=True)

class Model(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    brand_id = db.Column(db.Integer, db.ForeignKey('brand.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    purchases = db.relationship('Purchase', backref='model', lazy=True)
    sales = db.relationship('Sale', backref='model', lazy=True)

class Supplier(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact_info = db.Column(db.Text)
    address = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    purchases = db.relationship('Purchase', backref='supplier', lazy=True)

class Purchase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    model_id = db.Column(db.Integer, db.ForeignKey('model.id'), nullable=False)
    inventory_type = db.Column(db.String(10), nullable=False)  # 'new' or 'used'
    quantity = db.Column(db.Integer, nullable=False)
    purchase_price = db.Column(db.Float, nullable=False)
    imei_numbers = db.Column(db.Text, nullable=False)  # JSON string of IMEI numbers
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'), nullable=True)
    bill_number = db.Column(db.String(50), nullable=True)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='purchases')
    
    # Payment tracking fields (similar to Sale model)
    payment_status = db.Column(db.String(20), default='paid')  # 'paid', 'pending', 'partial'
    paid_amount = db.Column(db.Float, default=0.0)
    due_amount = db.Column(db.Float, default=0.0)
    due_date = db.Column(db.DateTime, nullable=True)  # Optional due date for supplier credit

class Shop(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    owner_name = db.Column(db.String(100), nullable=False)
    contact_info = db.Column(db.Text)
    address = db.Column(db.Text)
    credit_limit = db.Column(db.Float, default=0.0)  # Maximum credit allowed
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    sales = db.relationship('Sale', backref='shop', lazy=True)
    payments = db.relationship('Payment', backref='shop', lazy=True)

class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    model_id = db.Column(db.Integer, db.ForeignKey('model.id'), nullable=False)
    imei_number = db.Column(db.Text, nullable=False)  # Supports any IMEI format/length
    sale_price = db.Column(db.Float, nullable=False)
    purchase_price = db.Column(db.Float, nullable=False)  # Auto-fetched from purchase
    profit = db.Column(db.Float, nullable=False)  # Calculated automatically
    inventory_type = db.Column(db.String(10), nullable=False)  # 'new' or 'used'
    customer_type = db.Column(db.String(20), nullable=False)  # 'individual' or 'shop'
    shop_id = db.Column(db.Integer, db.ForeignKey('shop.id'), nullable=True)  # NULL for individual
    customer_name = db.Column(db.String(100), nullable=True)  # For individual customers
    payment_status = db.Column(db.String(20), default='paid')  # 'paid', 'pending', 'partial'
    paid_amount = db.Column(db.Float, default=0.0)
    due_amount = db.Column(db.Float, default=0.0)
    due_date = db.Column(db.DateTime, nullable=True)  # Optional due date
    bill_number = db.Column(db.String(50), nullable=True)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='sales')

class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shop_id = db.Column(db.Integer, db.ForeignKey('shop.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_date = db.Column(db.DateTime, default=datetime.utcnow)
    payment_method = db.Column(db.String(50), default='cash')  # 'cash', 'bank_transfer', 'cheque'
    reference_number = db.Column(db.String(100), nullable=True)  # For bank transfers/cheques
    notes = db.Column(db.Text, nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='payments')

class Incentive(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    brand_id = db.Column(db.Integer, db.ForeignKey('brand.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    description = db.Column(db.String(200))
    month = db.Column(db.Integer, nullable=False)  # 1-12
    year = db.Column(db.Integer, nullable=False)
    date = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='incentives')

class SupplierPayment(db.Model):
    """Track payments made to suppliers for purchases"""
    id = db.Column(db.Integer, primary_key=True)
    supplier_id = db.Column(db.Integer, db.ForeignKey('supplier.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    payment_date = db.Column(db.DateTime, default=datetime.utcnow)
    payment_method = db.Column(db.String(50), default='cash')  # 'cash', 'bank_transfer', 'cheque'
    reference_number = db.Column(db.String(100), nullable=True)  # For bank transfers/cheques
    notes = db.Column(db.Text, nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user = db.relationship('User', backref='supplier_payments')

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Routes
@app.route('/')
def index():
    """Redirect to dashboard if logged in, otherwise to login"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard with profit tracking"""
    try:
        # Calculate total profit including incentives
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        # Get all sales profit (filtered by user)
        total_sales_profit = db.session.query(db.func.sum(Sale.profit)).filter(
            Sale.user_id == current_user.id
        ).scalar() or 0
        
        # Get monthly sales profit (filtered by user)
        monthly_sales_profit = db.session.query(db.func.sum(Sale.profit)).filter(
            Sale.user_id == current_user.id,
            db.extract('month', Sale.date) == current_month,
            db.extract('year', Sale.date) == current_year
        ).scalar() or 0
        
        # Get incentives (filtered by user)
        total_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
            Incentive.user_id == current_user.id
        ).scalar() or 0
        monthly_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
            Incentive.user_id == current_user.id,
            Incentive.month == current_month,
            Incentive.year == current_year
        ).scalar() or 0
        
        # Calculate total profits
        total_profit_with_incentives = total_sales_profit + total_incentives
        monthly_profit_with_incentives = monthly_sales_profit + monthly_incentives
        
        # Get recent transactions (filtered by user)
        recent_sales = Sale.query.filter_by(user_id=current_user.id).order_by(Sale.date.desc()).limit(5).all()
        recent_purchases = Purchase.query.filter_by(user_id=current_user.id).order_by(Purchase.date.desc()).limit(5).all()
        
        # Get inventory stats (filtered by user)
        total_devices_purchased = db.session.query(db.func.sum(Purchase.quantity)).filter(
            Purchase.user_id == current_user.id
        ).scalar() or 0
        total_devices_sold = Sale.query.filter_by(user_id=current_user.id).count() or 0
        available_devices = total_devices_purchased - total_devices_sold
        
        return render_template('dashboard.html',
                             total_profit=total_profit_with_incentives,
                             monthly_profit=monthly_profit_with_incentives,
                             total_sales_profit=total_sales_profit,
                             monthly_sales_profit=monthly_sales_profit,
                             total_incentives=total_incentives,
                             monthly_incentives=monthly_incentives,
                             recent_sales=recent_sales,
                             recent_purchases=recent_purchases,
                             total_devices_purchased=total_devices_purchased,
                             total_devices_sold=total_devices_sold,
                             available_devices=available_devices)
    except Exception as e:
        # Return a simple dashboard with default values if there's an error
        return render_template('dashboard.html',
                             total_profit=0,
                             monthly_profit=0,
                             total_sales_profit=0,
                             monthly_sales_profit=0,
                             total_incentives=0,
                             monthly_incentives=0,
                             recent_sales=[],
                             recent_purchases=[],
                             total_devices_purchased=0,
                             total_devices_sold=0,
                             available_devices=0)

@app.route('/inventory_type')
@login_required
def inventory_type():
    """Inventory type selector - Used or New"""
    return render_template('inventory_type.html')

@app.route('/brands')
@login_required  
def brands():
    """Brand management and selection"""
    brands = Brand.query.all()
    
    # Prepare models data with stock information (filtered by user)
    models_data = {}
    for brand in brands:
        models_data[brand.id] = []
        for model in brand.models:
            # Calculate stock for this model (filtered by user)
            total_stock = db.session.query(db.func.sum(Purchase.quantity)).filter(
                Purchase.model_id == model.id,
                Purchase.user_id == current_user.id
            ).scalar() or 0
            
            sold_stock = Sale.query.filter_by(
                model_id=model.id,
                user_id=current_user.id
            ).count()
            available_stock = total_stock - sold_stock
            
            models_data[brand.id].append({
                'id': model.id,
                'name': model.name,
                'total_stock': total_stock,
                'available_stock': available_stock,
                'sold_stock': sold_stock
            })
    
    return render_template('brands.html', brands=brands, models_data=models_data)

@app.route('/add_brand', methods=['POST'])
@login_required
def add_brand():
    name = request.form.get('name')
    if name:
        brand = Brand(name=name)
        db.session.add(brand)
        db.session.commit()
        flash('Brand added successfully!', 'success')
    return redirect(url_for('brands'))

# @app.route('/models/<int:brand_id>')
# @login_required
# def models(brand_id):
#     """Model selection for a specific brand"""
#     brand = Brand.query.get_or_404(brand_id)
#     models = Model.query.filter_by(brand_id=brand_id).all()
#     return render_template('models.html', brand=brand, models=models)

@app.route('/add_model', methods=['POST'])
@login_required
def add_model():
    brand_id = request.form.get('brand_id')
    name = request.form.get('name')
    if brand_id and name:
        model = Model(brand_id=brand_id, name=name)
        db.session.add(model)
        db.session.commit()
        flash('Model added successfully!', 'success')
    
    # Always redirect to brands page
    return redirect(url_for('brands'))

@app.route('/purchase/<int:model_id>')
@login_required
def new_purchase(model_id):
    """New purchase entry form"""
    model = Model.query.get_or_404(model_id)
    return render_template('new_purchase.html', model=model)

@app.route('/add_purchase', methods=['POST'])
@login_required
def add_purchase():
    """Add new purchase with payment tracking"""
    model_id = request.form.get('model_id')
    inventory_type = request.form.get('inventory_type')
    quantity = int(request.form.get('quantity'))
    purchase_price = float(request.form.get('purchase_price'))
    imei_numbers = request.form.get('imei_numbers')  # Textarea with IMEIs
    supplier_id = request.form.get('supplier_id')
    bill_number = request.form.get('bill_number')
    
    # Payment tracking fields
    paid_amount = float(request.form.get('paid_amount') or 0)
    due_date_str = request.form.get('due_date')
    
    # Process IMEI numbers (split by newlines and clean)
    imei_list = [imei.strip() for imei in imei_numbers.split('\n') if imei.strip()]
    
    if len(imei_list) != quantity:
        flash(f'Number of IMEIs ({len(imei_list)}) must match quantity ({quantity})', 'error')
        return redirect(url_for('new_purchase', model_id=model_id))
    
    # Basic validation - just check if IMEIs are provided
    if not imei_list:
        flash('Please provide at least one IMEI number', 'error')
        return redirect(url_for('new_purchase', model_id=model_id))
    
    # Calculate total purchase amount and due amount
    total_amount = purchase_price * quantity
    due_amount = total_amount - paid_amount
    
    # Set payment status based on amount paid
    if paid_amount >= total_amount:
        payment_status = 'paid'
        due_amount = 0
    elif paid_amount > 0:
        payment_status = 'partial'
    else:
        payment_status = 'pending'
    
    # Parse due date if provided
    due_date = None
    if due_date_str:
        try:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
        except ValueError:
            due_date = None
    
    # Generate bill number if not provided
    if not bill_number:
        bill_number = f"PURCHASE-{datetime.now().strftime('%Y%m%d')}-{Purchase.query.count() + 1:04d}"
    
    # Save purchase with payment tracking
    purchase = Purchase(
        model_id=model_id,
        inventory_type=inventory_type,
        quantity=quantity,
        purchase_price=purchase_price,
        imei_numbers=json.dumps(imei_list),
        supplier_id=supplier_id if supplier_id else None,
        bill_number=bill_number,
        payment_status=payment_status,
        paid_amount=paid_amount,
        due_amount=due_amount,
        due_date=due_date,
        user_id=current_user.id
    )
    db.session.add(purchase)
    db.session.commit()
    
    # Create supplier payment record if any amount was paid
    if paid_amount > 0 and supplier_id:
        supplier_payment = SupplierPayment(
            supplier_id=supplier_id,
            amount=paid_amount,
            payment_method='cash',  # Default method
            notes=f'Payment for purchase bill: {bill_number}',
            user_id=current_user.id
        )
        db.session.add(supplier_payment)
        db.session.commit()
    
    flash(f'Purchase added successfully! Quantity: {quantity}, Bill: {bill_number}, Payment: {payment_status}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/add_supplier_payment/<int:purchase_id>', methods=['GET', 'POST'])
@login_required
def add_supplier_payment(purchase_id):
    """Add payment to supplier for existing purchase"""
    purchase = Purchase.query.get_or_404(purchase_id)
    
    # Ensure user owns this purchase
    if purchase.user_id != current_user.id:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        amount = float(request.form.get('amount'))
        payment_method = request.form.get('payment_method')
        reference_number = request.form.get('reference_number')
        notes = request.form.get('notes')
        
        if amount <= 0:
            flash('Payment amount must be greater than 0', 'error')
            return redirect(url_for('add_supplier_payment', purchase_id=purchase_id))
        
        # Create supplier payment record
        supplier_payment = SupplierPayment(
            supplier_id=purchase.supplier_id,
            amount=amount,
            payment_method=payment_method,
            reference_number=reference_number,
            notes=notes,
            user_id=current_user.id
        )
        db.session.add(supplier_payment)
        
        # Update purchase payment status
        purchase.paid_amount += amount
        purchase.due_amount = max(0, purchase.due_amount - amount)
        
        # Update payment status
        if purchase.due_amount == 0:
            purchase.payment_status = 'paid'
        else:
            purchase.payment_status = 'partial'
        
        db.session.commit()
        
        flash(f'Payment of PKR {amount:,.2f} added successfully!', 'success')
        return redirect(url_for('dashboard'))
    
    # GET request - show payment form
    return render_template('add_supplier_payment.html', purchase=purchase)

@app.route('/sale')
@login_required
def sale_module():
    """Sale module - search by IMEI"""
    return render_template('sale_module.html')

@app.route('/search_imei', methods=['POST'])
@login_required
def search_imei():
    """Search for device by IMEI"""
    imei = request.form.get('imei')
    
    if not imei:
        flash('Please enter an IMEI number', 'error')
        return redirect(url_for('sale_module'))
    
    # Allow any IMEI format - no validation constraints
    
    # Find the purchase containing this IMEI
    purchases = Purchase.query.all()
    found_purchase = None
    
    for purchase in purchases:
        imei_list = json.loads(purchase.imei_numbers)
        if imei in imei_list:
            # Check if already sold
            existing_sale = Sale.query.filter_by(imei_number=imei).first()
            if existing_sale:
                flash('This IMEI has already been sold!', 'error')
                return redirect(url_for('sale_module'))
            
            found_purchase = purchase
            break
    
    if not found_purchase:
        flash('IMEI not found in inventory!', 'error')
        return redirect(url_for('sale_module'))
    
    shops = Shop.query.all()
    return render_template('sale_form.html', purchase=found_purchase, imei=imei, shops=shops)

@app.route('/search_multiple_imei', methods=['POST'])
@login_required
def search_multiple_imei():
    """Search for multiple devices by IMEI"""
    imei_numbers = request.form.get('imei_numbers')
    
    if not imei_numbers:
        flash('Please enter IMEI numbers', 'error')
        return redirect(url_for('sale_module'))
    
    # Process IMEI numbers (split by newlines and clean)
    imei_list = [imei.strip() for imei in imei_numbers.split('\n') if imei.strip()]
    
    if not imei_list:
        flash('Please enter valid IMEI numbers', 'error')
        return redirect(url_for('sale_module'))
    
    # Find devices for each IMEI
    found_devices = []
    not_found_imeis = []
    already_sold_imeis = []
    
    for imei in imei_list:
        # Find the purchase containing this IMEI
        purchases = Purchase.query.filter_by(user_id=current_user.id).all()
        found_purchase = None
        
        for purchase in purchases:
            try:
                purchase_imei_list = json.loads(purchase.imei_numbers)
                if imei in purchase_imei_list:
                    # Check if already sold
                    existing_sale = Sale.query.filter_by(imei_number=imei, user_id=current_user.id).first()
                    if existing_sale:
                        already_sold_imeis.append(imei)
                        break
                    
                    found_purchase = purchase
                    break
            except (json.JSONDecodeError, TypeError):
                continue
        
        if found_purchase and imei not in already_sold_imeis:
            found_devices.append({
                'imei': imei,
                'purchase': found_purchase,
                'model': found_purchase.model,
                'brand': found_purchase.model.brand if found_purchase.model else None
            })
        elif imei not in already_sold_imeis:
            not_found_imeis.append(imei)
    
    # Check for errors
    error_messages = []
    if not_found_imeis:
        error_messages.append(f'IMEIs not found in inventory: {", ".join(not_found_imeis)}')
    if already_sold_imeis:
        error_messages.append(f'IMEIs already sold: {", ".join(already_sold_imeis)}')
    
    if error_messages:
        for msg in error_messages:
            flash(msg, 'error')
        return redirect(url_for('sale_module'))
    
    if not found_devices:
        flash('No valid devices found for sale', 'error')
        return redirect(url_for('sale_module'))
    
    shops = Shop.query.all()
    return render_template('multiple_sale_form.html', devices=found_devices, shops=shops)

@app.route('/add_multiple_sale', methods=['POST'])
@login_required
def add_multiple_sale():
    """Process multiple device sale"""
    try:
        customer_type = request.form.get('customer_type')
        pricing_method = request.form.get('pricing_method')
        
        if not customer_type or not pricing_method:
            flash('Please fill in all required fields', 'error')
            return redirect(url_for('sale_module'))
        
        # Get customer details
        shop_id = None
        customer_name = None
        due_date = None
        total_paid_amount = 0
        
        if customer_type == 'individual':
            customer_name = request.form.get('customer_name')
            total_paid_amount = float(request.form.get('total_paid_amount') or 0)
        elif customer_type == 'shop':
            shop_id = request.form.get('shop_id')
            due_date_str = request.form.get('due_date')
            if due_date_str:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
            total_paid_amount = float(request.form.get('paid_amount') or 0)
        
        # Process devices
        devices_data = []
        total_sale_amount = 0
        device_index = 0
        
        while True:
            model_id_key = f'devices[{device_index}][model_id]'
            if model_id_key not in request.form:
                break
                
            model_id = request.form.get(model_id_key)
            imei = request.form.get(f'devices[{device_index}][imei]')
            purchase_price = float(request.form.get(f'devices[{device_index}][purchase_price]'))
            inventory_type = request.form.get(f'devices[{device_index}][inventory_type]')
            
            # Get sale price based on pricing method
            if pricing_method == 'common':
                sale_price = float(request.form.get('common_sale_price'))
            else:
                sale_price = float(request.form.get(f'devices[{device_index}][sale_price]'))
            
            devices_data.append({
                'model_id': model_id,
                'imei': imei,
                'purchase_price': purchase_price,
                'sale_price': sale_price,
                'inventory_type': inventory_type,
                'profit': sale_price - purchase_price
            })
            
            total_sale_amount += sale_price
            device_index += 1
        
        if not devices_data:
            flash('No devices found to process', 'error')
            return redirect(url_for('sale_module'))
        
        # Generate bill number for the batch
        bill_number = f"BILL-{datetime.now().strftime('%Y%m%d')}-{Sale.query.count() + 1:04d}"
        
        # Create sales records
        created_sales = []
        for device in devices_data:
            # Calculate payment details for this device
            device_paid_amount = 0
            device_due_amount = device['sale_price']
            payment_status = 'paid'
            
            if customer_type == 'shop':
                # Distribute paid amount proportionally
                if total_paid_amount > 0:
                    proportion = device['sale_price'] / total_sale_amount
                    device_paid_amount = total_paid_amount * proportion
                    device_due_amount = device['sale_price'] - device_paid_amount
                
                if device_paid_amount >= device['sale_price']:
                    payment_status = 'paid'
                    device_due_amount = 0
                elif device_paid_amount > 0:
                    payment_status = 'partial'
                else:
                    payment_status = 'pending'
            else:
                # Individual customers - distribute paid amount proportionally
                if total_paid_amount > 0:
                    proportion = device['sale_price'] / total_sale_amount
                    device_paid_amount = total_paid_amount * proportion
                    device_due_amount = device['sale_price'] - device_paid_amount
                else:
                    device_paid_amount = 0
                    device_due_amount = device['sale_price']
                
                # Set payment status based on amount paid
                if device_paid_amount >= device['sale_price']:
                    payment_status = 'paid'
                    device_due_amount = 0
                elif device_paid_amount > 0:
                    payment_status = 'partial'
                else:
                    payment_status = 'pending'
            
            sale = Sale(
                model_id=device['model_id'],
                imei_number=device['imei'],
                sale_price=device['sale_price'],
                purchase_price=device['purchase_price'],
                profit=device['profit'],
                inventory_type=device['inventory_type'],
                customer_type=customer_type,
                shop_id=shop_id,
                customer_name=customer_name,
                payment_status=payment_status,
                paid_amount=device_paid_amount,
                due_amount=device_due_amount,
                due_date=due_date,
                bill_number=bill_number,
                user_id=current_user.id
            )
            
            db.session.add(sale)
            created_sales.append(sale)
        
        db.session.commit()
        
        # Create payment record for shop sales
        if customer_type == 'shop' and total_paid_amount > 0:
            payment = Payment(
                shop_id=shop_id,
                amount=total_paid_amount,
                payment_date=datetime.utcnow(),
                user_id=current_user.id
            )
            db.session.add(payment)
            db.session.commit()
        
        flash(f'Multiple sale completed successfully! {len(devices_data)} devices sold. Bill: {bill_number}', 'success')
        
        # Redirect to the first sale's bill (they all have the same bill number)
        return redirect(url_for('bill', sale_id=created_sales[0].id))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error processing multiple sale: {str(e)}', 'error')
        return redirect(url_for('sale_module'))

@app.route('/add_sale', methods=['POST'])
@login_required
def add_sale():
    """Process sale"""
    model_id = request.form.get('model_id')
    imei_number = request.form.get('imei_number')
    sale_price = float(request.form.get('sale_price'))
    purchase_price = float(request.form.get('purchase_price'))
    inventory_type = request.form.get('inventory_type')
    customer_type = request.form.get('customer_type')
    
    profit = sale_price - purchase_price
    
    # Initialize payment tracking
    payment_status = 'paid'
    paid_amount = sale_price
    due_amount = 0
    shop_id = None
    customer_name = None
    due_date = None
    
    if customer_type == 'individual':
        customer_name = request.form.get('customer_name')
        # Get amount paid by individual customer
        paid_amount = float(request.form.get('paid_amount') or 0)
        
        # Calculate due amount
        due_amount = sale_price - paid_amount
        
        # Set payment status based on amount paid
        if paid_amount >= sale_price:
            payment_status = 'paid'
            due_amount = 0
        elif paid_amount > 0:
            payment_status = 'partial'
        else:
            payment_status = 'pending'
    elif customer_type == 'shop':
        shop_id = request.form.get('shop_id')
        due_date_str = request.form.get('due_date')
        
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
        
        # Get amount paid by shop (if any)
        paid_amount = float(request.form.get('paid_amount') or 0)
        
        # Calculate due amount
        due_amount = sale_price - paid_amount
        
        # Set payment status based on amount paid
        if paid_amount >= sale_price:
            payment_status = 'paid'
            due_amount = 0
        elif paid_amount > 0:
            payment_status = 'partial'
        else:
            payment_status = 'pending'
    
    # Generate bill number
    bill_number = f"BILL-{datetime.now().strftime('%Y%m%d')}-{Sale.query.count() + 1:04d}"
    
    # Create sale record
    sale = Sale(
        model_id=model_id,
        imei_number=imei_number,
        sale_price=sale_price,
        purchase_price=purchase_price,
        profit=profit,
        inventory_type=inventory_type,
        customer_type=customer_type,
        shop_id=shop_id,
        customer_name=customer_name,
        payment_status=payment_status,
        paid_amount=paid_amount,
        due_amount=due_amount,
        due_date=due_date,
        bill_number=bill_number,
        user_id=current_user.id
    )
    
    db.session.add(sale)
    db.session.commit()
    
    flash('Sale completed successfully!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/add_payment/<int:sale_id>', methods=['GET', 'POST'])
@login_required
def add_payment(sale_id):
    """Add additional payment to an existing sale"""
    sale = Sale.query.filter_by(id=sale_id, user_id=current_user.id).first()
    
    if not sale:
        flash('Sale not found!', 'error')
        return redirect(url_for('transactions'))
    
    if request.method == 'POST':
        additional_payment = float(request.form.get('additional_payment') or 0)
        
        if additional_payment <= 0:
            flash('Please enter a valid payment amount', 'error')
            return redirect(url_for('add_payment', sale_id=sale_id))
        
        # Update payment details
        sale.paid_amount += additional_payment
        sale.due_amount = max(0, sale.sale_price - sale.paid_amount)
        
        # Update payment status
        if sale.paid_amount >= sale.sale_price:
            sale.payment_status = 'paid'
            sale.due_amount = 0
        else:
            sale.payment_status = 'partial'
        
        db.session.commit()
        
        flash(f'Payment of PKR {additional_payment:,.2f} added successfully!', 'success')
        return redirect(url_for('transactions'))
    
    return render_template('add_payment.html', sale=sale)

@app.route('/profits')
@login_required
def profits():
    """Password-protected profits screen"""
    return render_template('profits.html')

@app.route('/verify_profit_password', methods=['POST'])
@login_required
def verify_profit_password():
    """Verify password for profits screen"""
    password = request.form.get('password')
    
    if check_password_hash(current_user.password_hash, password):
        # Store verification in session
        session['profit_verified'] = True
        return redirect(url_for('profits_detailed'))
    else:
        flash('Incorrect password!', 'error')
        return redirect(url_for('profits'))

@app.route('/profits_detailed')
@login_required
def profits_detailed():
    """Detailed profits screen (password protected)"""
    if not session.get('profit_verified'):
        return redirect(url_for('profits'))
    
    try:
        # Calculate detailed profit information
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        # Get all sales profit (filtered by user)
        total_sales_profit = db.session.query(db.func.sum(Sale.profit)).filter(
            Sale.user_id == current_user.id
        ).scalar() or 0
        
        # Get monthly sales profit (filtered by user)
        monthly_sales_profit = db.session.query(db.func.sum(Sale.profit)).filter(
            Sale.user_id == current_user.id,
            db.extract('month', Sale.date) == current_month,
            db.extract('year', Sale.date) == current_year
        ).scalar() or 0
        
        # Get incentives (filtered by user)
        total_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
            Incentive.user_id == current_user.id
        ).scalar() or 0
        monthly_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
            Incentive.user_id == current_user.id,
            Incentive.month == current_month,
            Incentive.year == current_year
        ).scalar() or 0
        
        # Calculate total profits
        total_profit_with_incentives = total_sales_profit + total_incentives
        monthly_profit_with_incentives = monthly_sales_profit + monthly_incentives
        
        # Get profit breakdown by brand
        brand_profits = db.session.query(
            Brand.name,
            db.func.sum(Sale.profit).label('total_profit'),
            db.func.count(Sale.id).label('total_sales')
        ).select_from(Brand).join(Sale, Brand.id == Sale.model_id).filter(
            Sale.user_id == current_user.id
        ).group_by(Brand.name).all()
        
        # Get profit breakdown by month
        monthly_breakdown = db.session.query(
            db.extract('month', Sale.date).label('month'),
            db.extract('year', Sale.date).label('year'),
            db.func.sum(Sale.profit).label('profit')
        ).filter(
            Sale.user_id == current_user.id
        ).group_by(
            db.extract('month', Sale.date),
            db.extract('year', Sale.date)
        ).order_by(
            db.extract('year', Sale.date).desc(),
            db.extract('month', Sale.date).desc()
        ).limit(12).all()
        
        return render_template('profits_detailed.html',
                             total_profit=total_profit_with_incentives,
                             monthly_profit=monthly_profit_with_incentives,
                             total_sales_profit=total_sales_profit,
                             monthly_sales_profit=monthly_sales_profit,
                             total_incentives=total_incentives,
                             monthly_incentives=monthly_incentives,
                             brand_profits=brand_profits,
                             monthly_breakdown=monthly_breakdown)
    except Exception as e:
        flash(f'Error loading profits: {str(e)}', 'error')
        return redirect(url_for('profits'))

@app.route('/transactions')
@login_required
def transactions():
    """Transactions history - sales and purchases grouped by bill number with optional model filtering"""
    try:
        # Get model filter if provided
        model_id = request.args.get('model_id', type=int)
        
        # Get all sales (filtered by user, no profit info)
        query = Sale.query.filter_by(user_id=current_user.id)
        
        # Apply model filter if specified
        if model_id:
            query = query.filter_by(model_id=model_id)
            flash(f'Showing transactions for selected model only. <a href="{url_for("transactions")}" class="alert-link">View all transactions</a>', 'info')
        
        all_sales = query.order_by(Sale.date.desc()).all()
        
        # Group sales by bill number
        sales_groups = {}
        for sale in all_sales:
            bill_key = sale.bill_number or f"single_{sale.id}"
            if bill_key not in sales_groups:
                sales_groups[bill_key] = {
                    'bill_number': sale.bill_number,
                    'date': sale.date,
                    'customer_type': sale.customer_type,
                    'customer_name': sale.customer_name,
                    'shop': sale.shop,
                    'payment_status': sale.payment_status,
                    'items': [],
                    'total_amount': 0,
                    'total_paid': 0,
                    'total_due': 0
                }
            
            sales_groups[bill_key]['items'].append(sale)
            sales_groups[bill_key]['total_amount'] += sale.sale_price
            sales_groups[bill_key]['total_paid'] += sale.paid_amount
            sales_groups[bill_key]['total_due'] += sale.due_amount
            
            # Update payment status for the group (use the most restrictive status)
            current_status = sales_groups[bill_key]['payment_status']
            if sale.payment_status == 'pending' or current_status == 'pending':
                sales_groups[bill_key]['payment_status'] = 'pending'
            elif sale.payment_status == 'partial' or current_status == 'partial':
                sales_groups[bill_key]['payment_status'] = 'partial'
        
        # Convert to list and sort by date
        sales_groups_list = list(sales_groups.values())
        sales_groups_list.sort(key=lambda x: x['date'], reverse=True)
        
        # Get all purchases (filtered by user) - already grouped by nature
        purchases = Purchase.query.filter_by(user_id=current_user.id).order_by(Purchase.date.desc()).all()
        
        return render_template('transactions.html', sales_groups=sales_groups_list, purchases=purchases, all_sales=all_sales, model_filter=model_id)
    except Exception as e:
        flash(f'Error loading transactions: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/search_device')
@login_required
def search_device():
    """Search device by IMEI (shows all history)"""
    return render_template('search_device.html')

@app.route('/search_device_result', methods=['GET', 'POST'])
@login_required
def search_device_result():
    """Search result for device by IMEI"""
    try:
        # Get IMEI from either GET or POST request
        if request.method == 'POST':
            imei = request.form.get('imei')
        else:
            imei = request.args.get('imei')
        
        if not imei:
            flash('Please enter an IMEI number', 'error')
            return redirect(url_for('search_device'))
        
        # Clean IMEI - remove any whitespace or special characters
        imei = imei.strip()
        
        # Validate IMEI format (9-15 digits)
        if not imei.isdigit() or len(imei) < 9 or len(imei) > 15:
            flash('IMEI must be 9-15 digits', 'error')
            return redirect(url_for('search_device'))
        
        print(f"Searching for IMEI: {imei}")  # Debug log
        
        # Find the purchase containing this IMEI
        purchases = Purchase.query.filter_by(user_id=current_user.id).all()
        print(f"Found {len(purchases)} purchases for user {current_user.id}")  # Debug log
        
        found_purchase = None
        found_imei_index = -1
        
        for purchase in purchases:
            try:
                imei_list = json.loads(purchase.imei_numbers)
                print(f"Purchase {purchase.id} has {len(imei_list)} IMEIs: {imei_list[:3]}...")  # Debug log
                
                if imei in imei_list:
                    found_purchase = purchase
                    found_imei_index = imei_list.index(imei)
                    print(f"Found IMEI {imei} in purchase {purchase.id} at index {found_imei_index}")  # Debug log
                    break
            except (json.JSONDecodeError, TypeError) as e:
                # Handle corrupted IMEI data
                print(f"Error parsing IMEI data for purchase {purchase.id}: {e}")
                continue
        
        if not found_purchase:
            # Try to find similar IMEIs for better user feedback
            similar_imeis = []
            for purchase in purchases:
                try:
                    imei_list = json.loads(purchase.imei_numbers)
                    for stored_imei in imei_list:
                        if stored_imei.startswith(imei[:5]) or stored_imei.endswith(imei[-5:]):
                            similar_imeis.append(stored_imei)
                except:
                    continue
            
            if similar_imeis:
                flash(f'IMEI {imei} not found! Similar IMEIs in your inventory: {", ".join(similar_imeis[:3])}', 'error')
            else:
                flash(f'IMEI {imei} not found in your inventory!', 'error')
            return redirect(url_for('search_device'))
        
        # Check if sold
        existing_sale = Sale.query.filter_by(
            imei_number=imei,
            user_id=current_user.id
        ).first()
        
        print(f"Sale found for IMEI {imei}: {existing_sale.id if existing_sale else 'None'}")  # Debug log
        
        # Get supplier info if available
        supplier = None
        if found_purchase.supplier_id:
            supplier = Supplier.query.get(found_purchase.supplier_id)
        
        # Get model and brand info safely
        model = None
        brand = None
        if found_purchase.model_id:
            model = Model.query.get(found_purchase.model_id)
            if model and model.brand_id:
                brand = Brand.query.get(model.brand_id)
        
        print(f"Rendering template with: purchase={found_purchase.id}, model={model.name if model else 'None'}, brand={brand.name if brand else 'None'}")  # Debug log
        
        return render_template('search_device_result.html',
                             purchase=found_purchase,
                             imei=imei,
                             imei_index=found_imei_index,
                             sale=existing_sale,
                             supplier=supplier,
                             model=model,
                             brand=brand)
                             
    except Exception as e:
        print(f"Error in search_device_result: {e}")
        import traceback
        traceback.print_exc()  # Print full stack trace
        flash(f'Error searching for device: {str(e)}. Please try again.', 'error')
        return redirect(url_for('search_device'))

@app.route('/add_supplier', methods=['POST'])
@login_required
def add_supplier():
    """Add new supplier"""
    name = request.form.get('name')
    contact_info = request.form.get('contact_info')
    address = request.form.get('address')
    
    if name:
        supplier = Supplier(
            name=name,
            contact_info=contact_info,
            address=address
        )
        db.session.add(supplier)
        db.session.commit()
        flash('Supplier added successfully!', 'success')
    
    return redirect(url_for('new_purchase', model_id=request.form.get('model_id')))

@app.route('/get_suppliers')
@login_required
def get_suppliers():
    """Get suppliers for dropdown"""
    suppliers = Supplier.query.all()
    return jsonify([{'id': s.id, 'name': s.name} for s in suppliers])

@app.route('/validate_imei', methods=['POST'])
@login_required
def validate_imei():
    """Validate IMEI - check if it can be added"""
    imei = request.form.get('imei')
    inventory_type = request.form.get('inventory_type')
    
    if not imei:
        return jsonify({'valid': False, 'message': 'IMEI is required'})
    
    # Allow any IMEI format - no validation constraints
    
    # Check if IMEI already exists in purchases
    purchases = Purchase.query.filter_by(user_id=current_user.id).all()
    for purchase in purchases:
        imei_list = json.loads(purchase.imei_numbers)
        if imei in imei_list:
            # Check if it was sold
            existing_sale = Sale.query.filter_by(
                imei_number=imei,
                user_id=current_user.id
            ).first()
            
            if existing_sale:
                # IMEI was sold, can be bought again
                return jsonify({
                    'valid': True, 
                    'message': 'IMEI was previously sold, can be purchased again',
                    'was_sold': True
                })
            else:
                # IMEI exists and not sold
                return jsonify({
                    'valid': False, 
                    'message': 'IMEI already exists in inventory and has not been sold',
                    'was_sold': False
                })
    
    # IMEI is new
    return jsonify({'valid': True, 'message': 'IMEI is new and can be added'})

@app.route('/generate_bill_number')
@login_required
def generate_bill_number():
    """Generate unique bill number"""
    bill_number = f"BILL-{datetime.now().strftime('%Y%m%d')}-{Sale.query.count() + 1:04d}"
    return jsonify({'bill_number': bill_number})

@app.route('/bill/<int:sale_id>')
@login_required
def bill(sale_id):
    """Generate bill for sale with grouped sales information"""
    sale = Sale.query.filter_by(id=sale_id, user_id=current_user.id).first_or_404()
    
    # Get all sales for grouped transaction display
    all_sales = []
    if sale.bill_number:
        # Find all sales with the same bill number
        all_sales = Sale.query.filter_by(
            bill_number=sale.bill_number, 
            user_id=current_user.id
        ).order_by(Sale.id).all()
    
    return render_template('bill.html', sale=sale, all_sales=all_sales)

@app.route('/business_report/<int:sale_id>')
@login_required
def business_report(sale_id):
    """Business report with profit details for business analysis"""
    # Require profit verification
    if not session.get('profit_verified'):
        return redirect(url_for('profits'))
    sale = Sale.query.filter_by(id=sale_id, user_id=current_user.id).first_or_404()
    return render_template('business_report.html', sale=sale)

@app.route('/incentives')
@login_required
def incentives():
    """Incentive management"""
    brands = Brand.query.all()
    incentives = Incentive.query.filter_by(user_id=current_user.id).order_by(Incentive.date.desc()).all()
    
    # Calculate incentive totals (filtered by user)
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    total_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
        Incentive.user_id == current_user.id
    ).scalar() or 0
    monthly_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
        Incentive.user_id == current_user.id,
        Incentive.month == current_month,
        Incentive.year == current_year
    ).scalar() or 0
    
    return render_template('incentives.html', 
                         brands=brands, 
                         incentives=incentives,
                         total_incentives=total_incentives,
                         monthly_incentives=monthly_incentives,
                         now=datetime.now())

@app.route('/shops')
@login_required
def shops():
    """Shop management"""
    try:
        shops = Shop.query.all()
        
        # Calculate financial data for each shop (filtered by user)
        for shop in shops:
            shop.total_sales = db.session.query(db.func.sum(Sale.sale_price)).filter(
                Sale.shop_id == shop.id,
                Sale.user_id == current_user.id
            ).scalar() or 0
            
            shop.total_paid = db.session.query(db.func.sum(Sale.paid_amount)).filter(
                Sale.shop_id == shop.id,
                Sale.user_id == current_user.id
            ).scalar() or 0
            
            shop.total_due = db.session.query(db.func.sum(Sale.due_amount)).filter(
                Sale.shop_id == shop.id,
                Sale.user_id == current_user.id
            ).scalar() or 0
        
        return render_template('shops.html', shops=shops)
    except Exception as e:
        # Return empty shops list if there's an error
        return render_template('shops.html', shops=[])

@app.route('/add_shop', methods=['POST'])
@login_required
def add_shop():
    """Add new shop"""
    name = request.form.get('name')
    owner_name = request.form.get('owner_name')
    contact_info = request.form.get('contact_info')
    address = request.form.get('address')
    
    shop = Shop(
        name=name,
        owner_name=owner_name,
        contact_info=contact_info,
        address=address
    )
    db.session.add(shop)
    db.session.commit()
    
    flash('Shop added successfully!', 'success')
    return redirect(url_for('shops'))

@app.route('/shop/<int:shop_id>')
@login_required
def shop_details(shop_id):
    """Shop details with payment history"""
    # Check if user has any sales with this shop
    shop = Shop.query.get_or_404(shop_id)
    user_has_sales = Sale.query.filter_by(shop_id=shop_id, user_id=current_user.id).first() is not None
    if not user_has_sales:
        flash('Access denied. You can only view shops you have sales with.', 'error')
        return redirect(url_for('shops'))
    
    # Calculate shop's financial summary (filtered by user)
    total_sales = db.session.query(db.func.sum(Sale.sale_price)).filter(
        Sale.shop_id == shop_id,
        Sale.user_id == current_user.id
    ).scalar() or 0
    
    total_paid = db.session.query(db.func.sum(Sale.paid_amount)).filter(
        Sale.shop_id == shop_id,
        Sale.user_id == current_user.id
    ).scalar() or 0
    
    total_due = db.session.query(db.func.sum(Sale.due_amount)).filter(
        Sale.shop_id == shop_id,
        Sale.user_id == current_user.id
    ).scalar() or 0
    
    # Get recent sales and payments (filtered by user)
    recent_sales = Sale.query.filter_by(
        shop_id=shop_id,
        user_id=current_user.id
    ).order_by(Sale.date.desc()).limit(10).all()
    recent_payments = Payment.query.filter_by(
        shop_id=shop_id,
        user_id=current_user.id
    ).order_by(Payment.payment_date.desc()).limit(10).all()
    
    return render_template('shop_details.html', 
                         shop=shop,
                         total_sales=total_sales,
                         total_paid=total_paid,
                         total_due=total_due,
                         recent_sales=recent_sales,
                         recent_payments=recent_payments)

@app.route('/add_shop_payment', methods=['POST'])
@login_required
def add_shop_payment():
    """Add payment for a shop"""
    shop_id = request.form.get('shop_id')
    amount = float(request.form.get('amount'))
    payment_method = request.form.get('payment_method')
    reference_number = request.form.get('reference_number')
    notes = request.form.get('notes')
    
    payment = Payment(
        shop_id=shop_id,
        amount=amount,
        payment_method=payment_method,
        reference_number=reference_number,
        notes=notes,
        user_id=current_user.id
    )
    db.session.add(payment)
    db.session.commit()
    
    # Update shop's due amounts
    shop = Shop.query.get(shop_id)
    if shop:
        # Find pending sales and apply payment
        pending_sales = Sale.query.filter_by(
            shop_id=shop_id, 
            payment_status='pending'
        ).order_by(Sale.date.asc()).all()
        
        remaining_amount = amount
        for sale in pending_sales:
            if remaining_amount <= 0:
                break
                
            if sale.due_amount > 0:
                payment_amount = min(remaining_amount, sale.due_amount)
                sale.paid_amount += payment_amount
                sale.due_amount -= payment_amount
                
                if sale.due_amount == 0:
                    sale.payment_status = 'paid'
                else:
                    sale.payment_status = 'partial'
                
                remaining_amount -= payment_amount
        
        db.session.commit()
    
    flash(f'Payment of PKR {amount:,.2f} added successfully!', 'success')
    return redirect(url_for('shop_details', shop_id=shop_id))

@app.route('/pending_payments')
@login_required
def pending_payments():
    """View all pending payments"""
    try:
        # Get all shops with pending payments (filtered by user)
        shops_with_pending = db.session.query(Shop).select_from(Shop).join(Sale, Shop.id == Sale.shop_id).filter(
            Sale.payment_status.in_(['pending', 'partial']),
            Sale.user_id == current_user.id
        ).distinct().all()
        
        # Calculate pending amounts for each shop (filtered by user)
        shop_pending_data = []
        for shop in shops_with_pending:
            total_due = db.session.query(db.func.sum(Sale.due_amount)).filter(
                Sale.shop_id == shop.id,
                Sale.user_id == current_user.id
            ).scalar() or 0
            
            overdue_sales = Sale.query.filter(
                Sale.shop_id == shop.id,
                Sale.user_id == current_user.id,
                Sale.due_amount > 0,
                Sale.due_date < datetime.now()
            ).count()
            
            shop_pending_data.append({
                'shop': shop,
                'total_due': total_due,
                'overdue_sales': overdue_sales
            })
        
        return render_template('pending_payments.html', shop_pending_data=shop_pending_data)
    except Exception as e:
        # Return empty data if there's an error
        return render_template('pending_payments.html', shop_pending_data=[])

@app.route('/add_incentive', methods=['POST'])
@login_required
def add_incentive():
    """Add brand incentive"""
    brand_id = request.form.get('brand_id')
    amount = float(request.form.get('amount'))
    description = request.form.get('description')
    month = int(request.form.get('month'))
    year = int(request.form.get('year'))
    
    incentive = Incentive(
        brand_id=brand_id,
        amount=amount,
        description=description,
        month=month,
        year=year,
        user_id=current_user.id
    )
    db.session.add(incentive)
    db.session.commit()
    
    flash('Incentive added successfully!', 'success')
    return redirect(url_for('incentives'))

@app.route('/ledger')
@login_required
def shop_ledger():
    """Shop ledger with all transactions"""
    # Require profit verification since ledger shows profit totals
    if not session.get('profit_verified'):
        return redirect(url_for('profits'))
    brand_filter = request.args.get('brand')
    model_filter = request.args.get('model')
    
    # Get all transactions (filtered by user)
    purchases = Purchase.query.filter_by(user_id=current_user.id)
    sales = Sale.query.filter_by(user_id=current_user.id)
    
    if brand_filter:
        purchases = purchases.join(Model).filter(Model.brand_id == brand_filter)
        sales = sales.join(Model).filter(Model.brand_id == brand_filter)
    
    if model_filter:
        purchases = purchases.filter(Purchase.model_id == model_filter)
        sales = sales.filter(Sale.model_id == model_filter)
    
    purchases = purchases.order_by(Purchase.date.desc()).all()
    sales = sales.order_by(Sale.date.desc()).all()
    
    brands = Brand.query.all()
    models = Model.query.all()
    
    # Calculate profit totals (filtered by user)
    total_sales_profit = db.session.query(db.func.sum(Sale.profit)).filter(
        Sale.user_id == current_user.id
    ).scalar() or 0
    total_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
        Incentive.user_id == current_user.id
    ).scalar() or 0
    
    return render_template('ledger.html', 
                         purchases=purchases, 
                         sales=sales, 
                         brands=brands, 
                         models=models,
                         selected_brand=brand_filter,
                         selected_model=model_filter,
                         total_sales_profit=total_sales_profit,
                         total_incentives=total_incentives)

@app.route('/export_report')
@login_required
def export_report():
    """Generate and export PDF report"""
    # Require profit verification
    if not session.get('profit_verified'):
        return redirect(url_for('profits'))
    # Create PDF
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Mobile Shop Inventory Report', ln=True, align='C')
    pdf.ln(10)
    
    # Summary
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    total_sales_profit = db.session.query(db.func.sum(Sale.profit)).filter(
        Sale.user_id == current_user.id
    ).scalar() or 0
    monthly_sales_profit = db.session.query(db.func.sum(Sale.profit)).filter(
        Sale.user_id == current_user.id,
        db.extract('month', Sale.date) == current_month,
        db.extract('year', Sale.date) == current_year
    ).scalar() or 0
    
    total_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
        Incentive.user_id == current_user.id
    ).scalar() or 0
    monthly_incentives = db.session.query(db.func.sum(Incentive.amount)).filter(
        Incentive.user_id == current_user.id,
        Incentive.month == current_month,
        Incentive.year == current_year
    ).scalar() or 0
    
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Profit Summary', ln=True)
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 8, f'Total Sales Profit: PKR {total_sales_profit:,.2f}', ln=True)
    pdf.cell(0, 8, f'Monthly Sales Profit: PKR {monthly_sales_profit:,.2f}', ln=True)
    pdf.cell(0, 8, f'Total Incentives: PKR {total_incentives:,.2f}', ln=True)
    pdf.cell(0, 8, f'Monthly Incentives: PKR {monthly_incentives:,.2f}', ln=True)
    pdf.cell(0, 8, f'Total Profit (including incentives): PKR {(total_sales_profit + total_incentives):,.2f}', ln=True)
    pdf.ln(10)
    
    # Recent Sales
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'Recent Sales', ln=True)
    pdf.set_font('Arial', '', 10)
    
    recent_sales = Sale.query.order_by(Sale.date.desc()).limit(10).all()
    for sale in recent_sales:
        pdf.cell(0, 6, f'{sale.date.strftime("%Y-%m-%d")} - {sale.model.brand.name} {sale.model.name} - IMEI: {sale.imei_number} - Profit: PKR {sale.profit:.2f}', ln=True)
    
    # Save PDF
    pdf_buffer = BytesIO()
    pdf_content = pdf.output(dest='S').encode('latin-1')
    pdf_buffer.write(pdf_content)
    pdf_buffer.seek(0)
    
    return send_file(pdf_buffer, 
                     as_attachment=True, 
                     download_name=f'inventory_report_{datetime.now().strftime("%Y%m%d")}.pdf',
                     mimetype='application/pdf')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password', 'error')
    
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        
        if User.query.filter_by(email=email).first():
            flash('Email already exists', 'error')
            return render_template('signup.html')
        
        user = User(
            name=name,
            email=email,
            password_hash=generate_password_hash(password)
        )
        db.session.add(user)
        db.session.commit()
        
        flash('Account created successfully!', 'success')
        return redirect(url_for('login'))
    
    return render_template('signup.html')

@app.route('/logout')
@login_required
def logout():
    # Clear profit verification session
    session.pop('profit_verified', None)
    logout_user()
    return redirect(url_for('login'))

@app.route('/inventory')
@login_required
def inventory():
    """Show all inventory items with complete details"""
    try:
        # Get all purchases for current user
        purchases = Purchase.query.filter_by(user_id=current_user.id).order_by(Purchase.date.desc()).all()
        
        # Process each purchase to get detailed information
        inventory_items = []
        for purchase in purchases:
            try:
                imei_list = json.loads(purchase.imei_numbers)
                
                # Get model and brand info
                model = Model.query.get(purchase.model_id) if purchase.model_id else None
                brand = Brand.query.get(model.brand_id) if model and model.brand_id else None
                
                # Get supplier info
                supplier = Supplier.query.get(purchase.supplier_id) if purchase.supplier_id else None
                
                # Check status for each IMEI
                for imei in imei_list:
                    # Check if sold
                    sale = Sale.query.filter_by(
                        imei_number=imei,
                        user_id=current_user.id
                    ).first()
                    
                    inventory_items.append({
                        'imei': imei,
                        'purchase': purchase,
                        'model': model,
                        'brand': brand,
                        'supplier': supplier,
                        'sale': sale,
                        'status': 'Sold' if sale else 'Available',
                        'purchase_date': purchase.date,
                        'sale_date': sale.date if sale else None,
                        'profit': sale.profit if sale else None
                    })
            except (json.JSONDecodeError, TypeError) as e:
                print(f"Error processing purchase {purchase.id}: {e}")
                continue
        
        # Sort by purchase date (newest first)
        inventory_items.sort(key=lambda x: x['purchase_date'], reverse=True)
        
        return render_template('inventory.html', inventory_items=inventory_items)
        
    except Exception as e:
        print(f"Error in inventory route: {e}")
        flash('Error loading inventory. Please try again.', 'error')
        return redirect(url_for('dashboard'))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True) 